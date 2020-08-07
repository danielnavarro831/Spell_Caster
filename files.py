from openpyxl import workbook
from openpyxl import load_workbook

class Files:
    def __init__(self):
        #Change this path to your file path
        self.doc_path = 'C:\\Users\\Daniel\\source\\repos\\Spell Caster\\Spell_Caster_Strings.xlsx'
        self.save_path = 'C:\\Users\\Daniel\\source\\repos\\Spell Caster\\Spell_Caster_Save_File.xlsx'

    def __delete__(self):
        del self

    def Continue(self):
        wb = load_workbook(self.save_path)
        sheets = wb.sheetnames
        GameState = wb[sheets[5]]
        if GameState.cell(row = 5, column = 2).value == True:
            return True
        else:
            return False

    def save(self, Player, Game_State, Scene):
        wb = load_workbook(self.save_path)
        sheets = wb.sheetnames
        Speakers = wb[sheets[0]]
        for row in Speakers['B2:B10']:
          for cell in row:
            cell.value = None
        Stats = wb[sheets[1]]
        for row in Stats['B2:D26']:
          for cell in row:
            cell.value = None
        Inventory = wb[sheets[2]]
        for row in Inventory['A2:B300']:
          for cell in row:
            cell.value = None
        Rooms = wb[sheets[3]]
        for row in Rooms['A2:E20']:
          for cell in row:
            cell.value = None
        Interactables = wb[sheets[4]]
        for row in Interactables['A2:G40']:
          for cell in row:
            cell.value = None
        GameState = wb[sheets[5]]
        for row in GameState['B2:B10']:
          for cell in row:
            cell.value = None
        #Save Speaker Names
        Speakers.cell(row = 2, column = 2).value = Player.Name["Value"]
        Speakers.cell(row = 3, column = 2).value = Player.Speakers["Friend"]
        Speakers.cell(row = 4, column = 2).value = Player.Speakers["Grandfather"]
        Speakers.cell(row = 5, column = 2).value = Player.Speakers["Agatha"]
        #Save Stats
        All_Stats = Player.Stats + Player.Hidden_Stats
        for a in range(len(All_Stats)):
            counter = 2
            loop = True
            if All_Stats[a]["Name"] != "Name":
                while loop == True:
                    if Stats.cell(row = counter, column = 1).value == All_Stats[a]["Name"]:
                        Stats.cell(row = counter, column = 2).value = All_Stats[a]["Value"]
                        if "Casts" in All_Stats[a]:
                            Stats.cell(row = counter, column = 3).value = All_Stats[a]["Casts"]
                        if "Max" in All_Stats[a]:
                            Stats.cell(row = counter, column = 4).value = All_Stats[a]["Max"]
                        loop = False
                    else:
                        counter += 1
            else:
                continue
        #Save Inventory
        counter = 2
        for b in Player.Inventory.keys():
            Inventory.cell(row = counter, column = 1).value = b
            Inventory.cell(row = counter, column = 2).value = Player.Inventory[b]
            counter += 1
        if len(Player.Recipes) > 0:
            for b in range(len(Player.Recipes)):
                text = Player.Recipes[b] + " Recipe"
                Inventory.cell(row = counter, column = 1).value = text
                Inventory.cell(row = counter, column = 2).value = 1
                counter += 1
        #Save Rooms
        counter = 2
        counter_2 = 2
        for c in range(len(Game_State.tower)):
            #Name
            Rooms.cell(row = counter, column = 1).value = Game_State.tower[c].Name
            #Solved
            Rooms.cell(row = counter, column = 2).value = Game_State.tower[c].Solved
            #Counter
            Rooms.cell(row = counter, column = 3).value = Game_State.tower[c].Counter
            #Description
            Rooms.cell(row = counter, column = 4).value = Game_State.tower[c].Description
            #Steps
            if Game_State.is_dungeon(Game_State.tower[c]) == True:
                Rooms.cell(row = counter, column = 5).value = Game_State.tower[c].Steps_Taken
        #Save Interactables
            if len(Game_State.tower[c].Interactables) > 0:
                for d in range(len(Game_State.tower[c].Interactables)):
                    #Floor
                    Interactables.cell(row = counter_2, column = 1).value = Game_State.tower[c].Floor
                    #Name
                    Interactables.cell(row = counter_2, column = 2).value = str(Game_State.tower[c].Interactables[d].Name)
                    #Solved
                    Interactables.cell(row = counter_2, column = 3).value = Game_State.tower[c].Interactables[d].Solved
                    #Key Name
                    Interactables.cell(row = counter_2, column = 4).value = str(Game_State.tower[c].Interactables[d].Keys[0]["Name"])
                    #Value
                    Interactables.cell(row = counter_2, column = 5).value = Game_State.tower[c].Interactables[d].Keys[0]["Value"]
                    #Commands
                    Interactables.cell(row = counter_2, column = 6).value = str(Game_State.tower[c].Interactables[d].Keys[0]["Terms"])
                    #Type
                    Interactables.cell(row = counter_2, column = 7).value = Game_State.tower[c].Interactables[d].Keys[0]["Type"]
                    counter_2 += 1
            counter += 1
        #Save Game State
        GameState.cell(row = 2, column = 2).value = Game_State.current_floor
        GameState.cell(row = 3, column = 2).value = Game_State.in_dungeon
        GameState.cell(row = 4, column = 2).value = Game_State.tutorial
        GameState.cell(row = 5, column = 2).value = True
        wb.save(self.save_path) 
        print(Scene.print_message(85, False, "Menu", {}))

    def load(self, Player, Game_State, Controller, Scene):
        wb = load_workbook(self.save_path)
        sheets = wb.sheetnames
        Speakers = wb[sheets[0]]
        Stats = wb[sheets[1]]
        Inventory = wb[sheets[2]]
        Rooms = wb[sheets[3]]
        Interactables = wb[sheets[4]]
        GameState = wb[sheets[5]]
        #Set Speakers
        Player.Name["Value"] = Speakers.cell(row = 2, column = 2).value
        for a in Player.Speakers.keys():
            counter = 1
            while Speakers.cell(row = counter, column = 1).value != a:
                counter += 1
                if Speakers.cell(row = counter, column = 1).value == a:
                    Player.Speakers[a] = Speakers.cell(row = counter, column = 2).value
                    break
        #Set Stats
        All_Stats = Player.Stats + Player.Hidden_Stats
        for a in range(len(All_Stats)):
            if All_Stats[a]["Name"] != "Name":
                counter = 1
                while All_Stats[a]["Name"] != Stats.cell(row = counter, column = 1).value:
                    counter += 1
                    if All_Stats[a]["Name"] == Stats.cell(row = counter, column = 1).value:
                        All_Stats[a]["Value"] = Stats.cell(row = counter, column = 2).value
                        if "Casts" in All_Stats[a]:
                            All_Stats[a]["Casts"] = Stats.cell(row = counter, column = 3).value
                        if "Max" in All_Stats[a]:
                            All_Stats[a]["Max"] = Stats.cell(row = counter, column = 4).value
                        if All_Stats[a]["Name"] == "HP" or All_Stats[a]["Name"] == "MP":
                            All_Stats[a]["Value"] = All_Stats[a]["Max"]
                        break
            else:
                continue
        #Set Inventory
        counter = 2
        while Inventory.cell(row = counter, column = 1).value:
            if Inventory.cell(row = counter, column = 2).value > 0:
                if "Recipe" not in Inventory.cell(row = counter, column = 1).value:
                    Player.Inventory[Inventory.cell(row = counter, column = 1).value] = Inventory.cell(row = counter, column = 2).value
                else:
                    text = Inventory.cell(row = counter, column = 1).value
                    text = text.split(" Recipe")
                    Player.Recipes.append(text[0])
            counter += 1
        #Set Rooms
        for a in range(len(Game_State.tower)):
            Game_State.tower[a].Interactables = []
            counter = 1
            while Game_State.tower[a].Name != Rooms.cell(row = counter, column = 1).value:
                counter += 1
                if Game_State.tower[a].Name == Rooms.cell(row = counter, column = 1).value:
                    #Set Solved Status
                    Game_State.tower[a].Solved = Rooms.cell(row = counter, column = 2).value
                    #Set Counter
                    Game_State.tower[a].Counter = Rooms.cell(row = counter, column = 3).value
                    #Set Description
                    Game_State.tower[a].Description = Rooms.cell(row = counter, column = 4).value
                    if Game_State.is_dungeon(Game_State.tower[a]) == True:
                        Game_State.tower[a].Steps_Taken = Rooms.cell(row = counter, column = 5).value
                    break
        #Set Interactables
        counter = 2
        while Interactables.cell(row = counter, column = 1).value:
            for b in range(len(Game_State.tower)):
                if Game_State.tower[b].Floor == Interactables.cell(row = counter, column = 1).value:
                    Name = Interactables.cell(row = counter, column = 2).value
                    if "," in Name:
                        Name = self.make_list(Name)
                    Solved = Interactables.cell(row = counter, column = 3).value
                    KeyName = Interactables.cell(row = counter, column = 4).value
                    if "[" in KeyName:
                        KeyName = self.make_list(KeyName)
                    Value = Interactables.cell(row = counter, column = 5).value
                    Commands = Interactables.cell(row = counter, column = 6).value
                    if "[" in Commands:
                        Commands = self.make_list(Commands)
                    Type = Interactables.cell(row = counter, column = 7).value
                    Game_State.tower[b].add_interactable(Name, Solved, KeyName, Value, Commands, Type)
                    break
            counter += 1
        #Set Game State
        Game_State.current_floor = GameState.cell(row = 2, column = 2).value
        Game_State.in_dungeon = GameState.cell(row = 3, column = 2).value
        Game_State.tutorial = GameState.cell(row = 4, column = 2).value
        print(Scene.print_message(91, False, "Menu", {}))
        #Start Game
        Floor = 0
        for a in range(len(Game_State.tower)):
            if Game_State.tower[a].Name == Game_State.current_floor:
                Floor = a
                break
        if Game_State.in_dungeon == False:
            Game_State.tower[Floor].enter_room(Player, Scene, Controller, Game_State)
        else:
            Game_State.tower[Floor].enter_dungeon(Player, Scene, Controller, Game_State)

    def make_list(self, Text):
        Text = Text.strip("[")
        Text = Text.strip("]")
        Text_List = Text.split(", ")
        for a in range(len(Text_List)):
            Text_List[a] = Text_List[a].strip("'")
        return Text_List

    def delete_save_data(self, Player, Controller, Scene):
        #Confirmation Prompt
        loop = True
        confirmed = False
        while loop == True:
            #Are you sure?
            print(Scene.print_message(96, False, "Menu", {}))
            response = Controller.get_response(Player, False)
            if Controller.bools["Yes"] == True:
                confirmed = True
                loop = False
                print(Scene.print_message(94, False, "Menu", {}))
            elif Controller.bools["No"] == True:
                confirmed = False
                loop = False
        if confirmed == True:
            wb = load_workbook(self.save_path)
            sheets = wb.sheetnames
            Speakers = wb[sheets[0]]
            for row in Speakers['B2:B10']:
              for cell in row:
                cell.value = None
            Stats = wb[sheets[1]]
            for row in Stats['B2:D26']:
              for cell in row:
                cell.value = None
            Inventory = wb[sheets[2]]
            for row in Inventory['A2:B300']:
              for cell in row:
                cell.value = None
            Rooms = wb[sheets[3]]
            for row in Rooms['A2:E20']:
              for cell in row:
                cell.value = None
            Interactables = wb[sheets[4]]
            for row in Interactables['A2:G40']:
              for cell in row:
                cell.value = None
            GameState = wb[sheets[5]]
            for row in GameState['B2:B10']:
              for cell in row:
                cell.value = None
            GameState.cell(row = 5, column = 2).value = False
            wb.save(self.save_path) 
            print(Scene.print_message(95, False, "Menu", {}))

    def rename(self, Key, Name, Scene):
        wb = load_workbook(self.save_path)
        sheets = wb.sheetnames
        Speakers = wb[sheets[0]]
        counter = 1
        while Speakers.cell(row = counter, column = 1).value != Key:
            counter += 1
            if Speakers.cell(row = counter, column = 1).value == Key:
                Speakers.cell(row = counter, column = 2).value = Name
        wb.save(self.save_path)
        print(Scene.print_message(93, False, "Menu", {}))

    def choose_rename(self, Player, Controller, Scene):
        loop = True
        choices = []
        Key = ""
        Rename = ""
        #Choose Character to rename
        while loop == True:
            choices = ["Player"]
            print(Scene.print_message(97, False, "Menu", {}))
            print(" * Player")
            for a in Player.Speakers.keys():
                print(" * " + a)
                choices.append(a)
            response = Controller.get_response(Player, False)
            response = response.title()
            if response in choices:
                loop = False
                Key = response
        #Choose new name
        loop = True
        while loop == True:
            #What would you like to rename this character to?
            print(Scene.print_message(98, False, "Menu", {}))
            response = Controller.get_response(Player, False)
            response = response.title()
            if len(response) > 0 and len(response) < 17:
                #Confirmation prompt
                Rename = response
                loop2 = True
                while loop2 == True:
                    #Rename {Key} to {Name}?
                    print(Scene.print_message(99, False, "Menu", {"{Key}": Key, "{Name}": Rename}))
                    response = Controller.get_response(Player, False)
                    if Controller.bools["Yes"] == True:
                        loop2 = False
                        loop = False
                        self.rename(Key, Rename, Scene)
                    elif Controller.bools["No"] == True:
                        loop2 = False
            elif len(response) > 16:
                print(Scene.print_message(100, False, "Menu", {}))
            elif len(response) < 1:
                print(Scene.print_message(101, False, "Menu", {}))

    def fix(self, Page, Row, Column, Value):
        wb = load_workbook(self.save_path)
        sheets = wb.sheetnames
        sheet = wb[sheets[Page]]
        sheet.cell(row = Row, column = Column).value = Value
        wb.save(self.save_path)
